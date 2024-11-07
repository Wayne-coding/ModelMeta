import pandas as pd
import mindspore
import mindspore as ms
import collections
import uuid
import copy
import datetime
import json
import os
import platform
import random
import sys
import time
from mindspore import JitConfig
import mindspore.context as context
from mindspore import export, load_checkpoint, load_param_into_net
from mindspore.rewrite import ScopedValue, NodeType
from mindspore.rewrite.node import Node, NodeManager
from numpy import ndarray
from openpyxl import Workbook
import mindspore.numpy as mnp
from mindspore import Tensor
from infoplus.MindSporeInfoPlus import mindsporeinfoplus
import torch
import torch.optim as optim
from mindspore import Tensor
from mindspore.rewrite import SymbolTree

import pickle
from mindspore_mutation.cargo import *
import torch.distributions as dist
import copy
import time
import json
import torch.fx as fx
from mindspore_mutation.MR_structure import *
from mindspore_mutation.cargo import match_rule,reflect_name,MCMC,compute_gpu_cpu
from mindspore_mutation.api_mutation import api_mutation
from mindspore_mutation.calculate_coverage import model2cov,find_layer_type
from mindspore_mutation.cargo import select_places,max_seed_model_api_times
import psutil
import sys
from openpyxl import Workbook
import os
import numpy as np
from mindspore import Tensor
from mindspore import dtype as mstype
import mindspore.context as context
from mindspore import save_checkpoint
import torch_mutation.config as pt_config
import mindspore_mutation.config as ms_config

from mindspore_mutation.handel_shape import handle_format
from mindspore_mutation import metrics
import gc


ms_device=ms_config.device
pt_device=pt_config.device
ms.set_context(mode=ms.PYNATIVE_MODE, device_target=ms_device)

MR_structure_name_list = ['UOC', 'PIOC', 'ABSOC_A', 'ABSOC_B']
MR_structures_map = {"UOC": UOC, "PIOC": PIOC, "ABSOC_A": ABSOC_A, "ABSOC_B": ABSOC_B} 
nlp_cargo = ["LSTM","FastText", "TextCNN", "SentimentNet", "GPT"]

deadcode_name_list=['Dense', 'SELayer', 'DenseLayer', 'Inception_A', 'PWDWPW_ResidualBlock', 'ResidualBlock', 'DropPath']



def run_mcmc_ms(seed_model, mutate_times,num_samples):
    localtime = time.strftime('%Y_%m_%d_%H_%M_%S', time.localtime())
    log_dict = {}
    dtypes = [mindspore.float32] if seed_model not in nlp_cargo else [mindspore.int32]
    MCMC_selector = MCMC()
    last_MR_structure_name = None
    last_reward = 0  
    log_dict = {}

    
    if isinstance(datasets_path_cargo[seed_model],list):
        data_0 = np.load(datasets_path_cargo[seed_model][0])
        data_1 = np.load(datasets_path_cargo[seed_model][1])
        samples_0 = np.random.choice(data_0.shape[0], num_samples, replace=False)
        samples_data_0 = data_0[samples_0] 

        samples_1 = np.random.choice(data_1.shape[0], num_samples, replace=False)
        samples_data_1 = data_1[samples_1] 
        data_selected_0 = Tensor(samples_data_0, dtype=mstype.float32 if seed_model in nlp_cargo else mstype.int32)
        data_selected_1 = Tensor(samples_data_1, dtype=mstype.float32 if seed_model in nlp_cargo else mstype.int32)
        data_selected = (data_selected_0, data_selected_1)
        data_npy = [data_selected_0.asnumpy(), data_selected_1.asnumpy()]

        npy_path = os.path.join("results", seed_model, str(localtime), 'data0_npy.npy')
        os.makedirs(os.path.dirname(npy_path), exist_ok=True)
        np.save(npy_path, data_npy[0])
        npy_path = os.path.join("results", seed_model, str(localtime), 'data1_npy.npy')
        os.makedirs(os.path.dirname(npy_path), exist_ok=True)
        np.save(npy_path, data_npy[1])
    else:
        data = np.load(datasets_path_cargo[seed_model])
        samples = np.random.choice(data.shape[0], num_samples, replace=False)
        samples_data = data[samples] 
        data_selected = Tensor(samples_data, dtype=mstype.int32 if seed_model in nlp_cargo else mstype.float32)
        data_npy = data_selected.asnumpy()


        npy_path = os.path.join("results", seed_model, str(localtime), 'data0_npy.npy')
        os.makedirs(os.path.dirname(npy_path), exist_ok=True)
        np.save(npy_path, data_npy)

    seed_model_net = get_model(seed_model)
    new_net = copy.deepcopy(seed_model_net)
    stree = SymbolTree.create(new_net)

    option_layers = []
    for name, child in new_net.cells_and_names():
        if not has_child_node(new_net, name) and not name == '' and not 'deadcode' in str(type(child)):
            if name.split("_")[0] not in MR_structure_name_list:
                option_layers.append((name, child, name, type(child)))

    
    
    original_outputs = handle_format(new_net(data_selected))
    new_outputs = original_outputs

    metrics_dict = dict()
    select_d_name = seed_model
    D = {seed_model: stree} 
    O = {seed_model: original_outputs} 
    N = {seed_model: seed_model_net} 
    R = {0:[0.0001, seed_model]} 
    MR_structure_selected_nums = {k: 0 for k in MR_structure_name_list}  
    seed_model_api_times=0 

    tar_set = set()
    tar_set_all = []
    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=1).value = 'name'
    sheet.cell(row=1, column=2).value = 'input_cov'
    sheet.cell(row=1, column=3).value = 'config_cov'
    sheet.cell(row=1, column=4).value = 'api_cov'
    sheet.cell(row=1, column=5).value = 'op_type_cov'
    sheet.cell(row=1, column=6).value = 'op_num_cov'
    sheet.cell(row=1, column=7).value = 'edge_cov'
    sheet.cell(row=1, column=8).value = 'inside'
    sheet.cell(row=1, column=9).value = 'outside'
    sheet.cell(row=1, column=10).value = 'distance'
    counter = 2
    counter_l = 2

    
    for n in range(mutate_times):
        print('-----------------------total_Mutate_time:%d start!-----------------------' % n)
        start_time=time.time()
        try:
            log_dict[n] = {}
            log_dict[n]['d_name'] = select_d_name

            
            selected_deadcode_name=random.choice(deadcode_name_list)

            
            selected_MR_structure_name = MCMC_selector.choose_mutator(last_MR_structure_name)
            selected_MR_structure = MCMC_selector.mutators[selected_MR_structure_name]
            selected_MR_structure.total += 1
            last_MR_structure_name = selected_MR_structure_name
            MR_structure_selected_nums[selected_MR_structure_name] += 1


            
            d_new_name = "{}-{}{}".format(seed_model, selected_MR_structure_name,MR_structure_selected_nums[selected_MR_structure_name])
            log_dict[n]['d_new_name'] = d_new_name

            
            if selected_deadcode_name in ('DropPath', 'Dense') and seed_model_api_times < max_seed_model_api_times(seed_model):  
                api_mutation_type = 'seed_model'
                seed_model_api_times+=1
            elif selected_deadcode_name not in ('DropPath', 'Dense') and seed_model_api_times < max_seed_model_api_times(seed_model): 
                api_mutation_type = random.choice(['seed_model', 'deadcode'])  
                if api_mutation_type=='seed_model':
                    seed_model_api_times += 1
            elif selected_deadcode_name not in ('DropPath', 'Dense') and seed_model_api_times >= max_seed_model_api_times(seed_model):  
                api_mutation_type = 'deadcode'
            else: 
                api_mutation_type = 'None'
                log_dict[n]['state'] = "Success:But no APIs available for mutation, so no API-level mutation was performed."
            



            
            nodedict = collections.OrderedDict()  
            hash_table = collections.defaultdict(int)  
            flag, nodedict = scan_node(stree, hash_table, nodedict)
            
            
            
            length=len(nodedict)
            print('length', length)
            print("mutate_type:", selected_MR_structure_name, ";  op_type:", selected_deadcode_name, ';  api_mutation_type:', api_mutation_type, flush=True)
            sys.setrecursionlimit(4000)
            def select_node(nodedict, recurive_depth=0):
                if recurive_depth >= 3500:
                    return None, None, None, None, recurive_depth
                subs_place, dep_places = \
                    select_places(range(0, length - 1), 5)
                if dep_places is None:
                    return select_node(nodedict, recurive_depth + 1)
                dep_places.sort(reverse=True)
                node_list = []
                for k, v in nodedict.items():
                    node_list.append(k)
                a = node_list[dep_places[-1]]
                b = node_list[dep_places[-2]]
                c = node_list[dep_places[-3]]
                d = node_list[dep_places[-4]]
                a = mindspore.rewrite.api.node.Node(a)
                b = mindspore.rewrite.api.node.Node(b)
                c = mindspore.rewrite.api.node.Node(c)
                d = mindspore.rewrite.api.node.Node(d)
                if not a._node.get_belong_symbol_tree() == b._node.get_belong_symbol_tree() == c._node.get_belong_symbol_tree() == d._node.get_belong_symbol_tree():
                    return select_node(nodedict, recurive_depth + 1)
                elif not (check_node(d) and check_node(c) and check_node(b) and check_node(a)):
                    return select_node(nodedict, recurive_depth + 1)
                elif selected_MR_structure_name == "PIOC" and (c.get_users()[0].get_node_type() == NodeType.Output or c.get_users()[0].get_node_type() == NodeType.Tree):
                    return select_node(nodedict, recurive_depth + 1)
                elif d.get_users()[0].get_node_type() == NodeType.Output or d.get_users()[0].get_node_type() == NodeType.Tree :
                    return select_node(nodedict, recurive_depth + 1)
                else:
                    log_dict[n]['subs_place'], log_dict[n]['dep_places'] = subs_place, dep_places
                    return a, b, c, d, recurive_depth

            start_time = time.time()
            aa, bb, cc, dd, recurive_depth = select_node(nodedict, 0)
            end_time = time.time()
            print("recurive depth:", recurive_depth, "select_node time:", end_time - start_time)
            if recurive_depth>=3500: 
                log_dict[n]['state'] = f"Failed:Cannot find suitable places！"
            if aa is None:
                print("mutate failed for Cannot find suitable places")
                
            
            
            
            add_module = MR_structures_map[selected_MR_structure_name](selected_deadcode_name, api_mutation_type, log_dict, n, LOG_FLAG=False)

            seat = 0
            if selected_MR_structure_name == "PIOC" and selected_deadcode_name in ["Dense", "Conv", "SELayer", "DenseLayer", "Inception_A",
                                                "PWDWPW_ResidualBlock", "ResidualBlock", "DropPath"]:
                
                if not (check_node(cc) and check_node(bb) and check_node(aa)):
                    log_dict[n]['state'] = "Failed:选择插入的节点位置不正确！"
                    print("Failed:选择插入的节点位置不正确！")
                    

                tree = cc.get_symbol_tree()  

                position = tree.after(cc)  
                next_node = cc.get_users()[0]  
                
                
                
                
                
                    
                if len(next_node.get_args()) > 1:
                    for idx, arg in enumerate(next_node.get_args()):
                        if arg == cc.get_targets()[0]:
                            seat = idx  
                            break
                        
                new_node = mindspore.rewrite.api.node.Node.create_call_cell(add_module,
                                                                            
                                                                            targets=[stree.unique_name("x")],
                                                                            name="{}_{}".format(selected_MR_structure_name,MR_structure_selected_nums[selected_MR_structure_name]),  
                                                                            args=ScopedValue.create_name_values(["aa", "bb", "cc"]))
                new_node.set_arg_by_node(0, aa) 
                new_node.set_arg_by_node(1, bb)
                new_node.set_arg_by_node(2, cc)
            else:  

                if not (check_node(dd) and check_node(cc) and check_node(bb) and check_node(aa)):
                    log_dict[n]['state'] = "Failed:选择插入的节点位置不正确！"
                    print("Failed:选择插入的节点位置不正确！")
                    
                tree = dd.get_symbol_tree()
                position = tree.after(dd)
                next_node = dd.get_users()[0]
                
                
                
                
                
                    
                if len(next_node.get_args()) > 1:
                    for idx, arg in enumerate(next_node.get_args()):
                        if arg == dd.get_targets()[0]:
                            seat = idx
                            break
                new_node = mindspore.rewrite.api.node.Node.create_call_cell(add_module,
                                                                            
                                                                            targets=[stree.unique_name("x")],
                                                                            name="{}_{}".format(selected_MR_structure_name,MR_structure_selected_nums[selected_MR_structure_name]),
                                                                            args=ScopedValue.create_name_values(["aa", "bb", "cc","dd"]))
                new_node.set_arg_by_node(0, aa)
                new_node.set_arg_by_node(1, bb)
                if selected_MR_structure_name == "UOC":
                    new_node.set_arg_by_node(2, cc)
                    
                    new_node.set_arg_by_node(3, dd)
                    
                else:
                    new_node.set_arg_by_node(2, dd)
                    new_node.set_arg_by_node(3, cc)
            tree.insert(position, new_node)
            next_node.set_arg_by_node(seat, new_node)
            D[d_new_name] = stree  
            
            new_net = stree.get_network()
            if api_mutation_type == 'seed_model':
                try:
                    new_net, stree, log_dict = api_mutation(new_net, option_layers, log_dict, n, LOG_FLAG=False) 
                    print(f"Success during api_mutation")
                except Exception as e:
                    print(f"Error during api_mutation: {e}")
                    log_dict[n]['state'] = f"Failed: api_mutation failed: {str(e)}"

            N[d_new_name] = new_net  
            new_outputs = new_net(data_selected) 
            
            
            
            new_output = handle_format(new_outputs)
            
            
            
            O[d_new_name] = copy.deepcopy(new_output)
            print('ChebyshevDistance:',metrics.ChebyshevDistance(original_outputs,new_output),';  MAEDistance:',metrics.MAEDistance(original_outputs,new_output))
            dist_chess = metrics.ChebyshevDistance(original_outputs,new_output)
            if new_output.shape!=original_outputs.shape:
                print('new_output.shape!=original_outputs.shape!')
                

        except Exception as e:
            print(e)
            log_dict[n]['state'] = f"Failed: Error during mutation: {str(e)}"

       
        
        with torch.no_grad():
            if ('state' in log_dict[n]) and ("Success" not in log_dict[n]['state']):  
                
                reward,done=-1,True
                
                d_probs = torch.distributions.Beta(torch.tensor([value[0] for value in R.values()]), torch.ones(len(R))).sample()
                select_d_name = R[torch.argmax(d_probs).item()][1]
                
                
                stree = D[select_d_name]
                
                metrics_dict[d_new_name]=["None"]*4
                
                
                next_output = O[select_d_name]
                
                
                
                
                formatted_data = O[select_d_name]
                
                
            else:
                done = False
                json_file_path= os.path.join("/data1/czx/SemTest_master/mindspore_mutation/results", seed_model, str(localtime),"model_json" , str(n) + ".json")
                os.makedirs(os.path.join("/data1/czx/SemTest_master/mindspore_mutation/results", seed_model, str(localtime),"model_json"), exist_ok=True)
                all_json_path="/data1/czx/SemTest_master/mindspore_mutation/ms_all_layer_info.json"
                api_config_pool_path = '/data1/czx/SemTest_master/mindspore_mutation/mindspore_api_config_pool.json'
                folder_path = '/data1/czx/SemTest_master/'
                
                input_data = mindsporeinfoplus.np_2_tensor([data_npy], dtypes)
                
                stree = stree 
                input_cov, config_cov, api_cov, op_type_cov, op_num_cov, edge_cov, inside, output_datas = model2cov(new_net,
                                                                                                        input_data,
                                                                                                        dtypes,
                                                                                                        json_file_path,
                                                                                                        all_json_path,
                                                                                                        api_config_pool_path,
                                                                                                        folder_path)
                reward= (input_cov+config_cov+api_cov)/3
                metrics_dict[d_new_name]=[input_cov,config_cov,api_cov,reward]
                R[len(R)]=[reward,d_new_name] 
                select_d_name=d_new_name
                
                
                next_output = O[d_new_name]
                
                formatted_data = O[d_new_name]
                

        selected_MR_structure.delta_bigger_than_zero = selected_MR_structure.delta_bigger_than_zero + 1 \
            if (reward - last_reward) > 0 else selected_MR_structure.delta_bigger_than_zero
        last_reward = reward

        end_time = time.time()
        elapsed_time = end_time - start_time  
        metrics_dict[d_new_name].append(elapsed_time)
        del formatted_data
        gc.collect()
        torch.cuda.empty_cache()

        
        if ('state' in log_dict[n]) and ("Success" not in log_dict[n]['state']):  
            log_dict[n]['select_d_name'] = select_d_name
        else:
            log_dict[n]['state']='Success!'
            log_dict[n]['select_deadcode'] = selected_deadcode_name
            log_dict[n]['selected_MR_structure'] = selected_MR_structure_name
            log_dict[n]['api_mutation_type(seed_model or deadcode)'] = api_mutation_type
            log_dict[n]['select_d_name'] = select_d_name
        
        dict_save_path = os.path.join("results", seed_model, str(localtime),"TORCH_LOG_DICT_" + str(ms_device).replace(':', '_') + ".json")
        os.makedirs(os.path.dirname(dict_save_path), exist_ok=True)
        with open(dict_save_path, 'w', encoding='utf-8') as file:
            json.dump(log_dict, file, ensure_ascii=False, indent=4)

        
        df = pd.DataFrame([(k, v[0], v[1],v[2],v[3],v[4]) for k, v in metrics_dict.items()],
                          columns=['New_Model_Name', 'Input_cov','Config_cov','Api_cov','Avg_cov','Elapsed_time'])
        save_path = os.path.join("results", seed_model, str(localtime),"METRICS_RESULTS_" + str(ms_device).replace(':', '_') + ".xlsx")
        df.to_excel(save_path, index=False)


        layer_type = find_layer_type(new_net)
        if len(set(layer_type)) > len(tar_set):  
            tar_set = set(layer_type)
        tar_set_all.append(set(layer_type))
        if 'input_cov' in dir():
            sheet.cell(row=counter, column=1).value = str(n)
            sheet.cell(row=counter, column=2).value = input_cov
            sheet.cell(row=counter, column=3).value = config_cov
            sheet.cell(row=counter, column=4).value = api_cov
            sheet.cell(row=counter, column=5).value = op_type_cov
            sheet.cell(row=counter, column=6).value = op_num_cov
            sheet.cell(row=counter, column=7).value = edge_cov
            sheet.cell(row=counter, column=8).value = inside
            sheet.cell(row=counter, column=10).value = dist_chess
            counter += 1
        index_1 = 2
        counter_l = 2
        if 'tar_set_all' in dir():
            for ii in tar_set_all:
                outer_div = len(tar_set - ii)  
                sheet.cell(row=counter_l, column=9).value = outer_div
                counter_l += 1
        data_path = os.path.join("results", seed_model, str(localtime),"METRICS_RESULTS_" + str(ms_device).replace(':', '_') + "ALL.xlsx")
        workbook.save(data_path)

        print('state',log_dict[n]['state'])

        print('-----------------------total_Mutate_time:%d ended!-----------------------' % n)
        
        new_outputs = next_output

        

